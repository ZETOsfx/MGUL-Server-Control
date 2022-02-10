import json
import xform
from openpyxl.styles import *
from openpyxl.utils.cell import get_column_letter

from progress.bar import IncrementalBar
from PyQt5 import QtCore, QtWidgets


# окно выбора файла
app = QtWidgets.QApplication([])
fileName = QtWidgets.QFileDialog.getOpenFileName()[0]
# progress bar
_bar = [1, 2, 3, 4, 5, 6]
bar = IncrementalBar('SERVER STATE: ', max=len(_bar))

# STYLE PREFABS
# 1. fill
_ok = PatternFill(start_color="d9ead3",
                  end_color="d9ead3", fill_type="solid")   # зелёный
_warn = PatternFill(start_color="fff2cc",
                    end_color="fff2cc", fill_type="solid")  # жёлтый
_err = PatternFill(start_color="f4cccc",
                   end_color="f4cccc", fill_type="solid")   # красный
# 2. font
_bold = Font(name='Calibri', bold=True, size=14, color="000000")
# 3. border
_selfbord = Border(left=Side(style='medium'), right=Side(
    style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))  # Линия границы ячейки


book = xform.create_excel()   # таблица для вывода


def toFixed(numObj, digits=0):  # ограничение знаков после запятой
    return f"{numObj:.{digits}f}"


def startValues():          # начальные значения переменных для обработки
    global i5               # записи во время, где число минут кратно 5
    global ir               # все остальные записи с рандомным временем
    global last_min     # последняя минута
    global twice_5      # проверка на избыточные граничные записи
    global twice        # проверка на дублирование записей в целом
    global extra        # флаг допуска к условию пересчета и печати
    global skip         # флаг обхода дополнительной проверки
    global GO_IN        # флаг последнего вхождения для печати оставшейся информации
    global am_pm        # перевод формата для перехода между диапазонами
    global interval     # не ну это ты понял

    i5 = 0              # записи во время, где число минут кратно 5
    ir = 0              # все остальные записи с рандомным временем
    last_min = 0
    twice_5 = False     # проверка на избыточные граничные записи
    twice = False       # проверка на дублирование записей в целом
    extra = False       # флаг допуска к условию пересчета и печати
    skip = False        # флаг обхода дополнительной проверки
    GO_IN = False       # флаг последнего вхождения для печати оставшейся информации
    am_pm = False       # перевод формата для перехода между диапазонами
    interval = 3


def Processing(_dev, _ser):  # основной цикл обработки (универсальный для каждого сервера)
    global i5               # записи во время, где число минут кратно 5
    global ir               # все остальные записи с рандомным временем
    global last_min     # последняя минута
    global twice_5      # проверка на избыточные граничные записи
    global twice        # проверка на дублирование записей в целом
    global extra        # флаг допуска к условию пересчета и печати
    global skip         # флаг обхода дополнительной проверки
    global GO_IN        # флаг последнего вхождения для печати оставшейся информации
    global am_pm        # перевод формата для перехода между диапазонами
    global interval     # не ну это ты понял

    for key in data:
        # chek selected device (ИМЕННО НА ВЫБРАННЫЙ ДЕВАЙС)
        uName = data[str(key)]['uName']
        key_copy = int(key) + 1
        # проверка наличия след. записи (конец файла .json)
        try:
            data[str(key_copy)]['uName']
        except:
            # флаги всех обходов для последнего вхождения и печати
            GO_IN = True
            twice_5 = False
            extra = True
            skip = True
        # определение записей необходимого девайса
        if (uName == device[_dev] and data[str(key)]['serial'] == _ser) or GO_IN:

            # переключаем лист на нужный для каждого сревера
            if uName == device['dbrobo']:
                book.active = 0
            elif uName == device['webrobo']:
                book.active = 1
            elif uName == device['dokuwiki']:
                book.active = 2
            elif uName == device['sev']:
                if data[str(key)]['serial'] == '01':
                    book.active = 3
                elif data[str(key)]['serial'] == '02':
                    book.active = 4
                elif data[str(key)]['serial'] == '03':
                    book.active = 5

            # время записи (не для последнего вхождения)
            if not GO_IN:
                hour = int(data[str(key)]['Date'][11] +
                           data[str(key)]['Date'][12])
                min = int(data[str(key)]['Date'][14] +
                          data[str(key)]['Date'][15])
                minNext = int(data[str(int(key) + 1)]['Date']
                              [14] + data[str(int(key) + 1)]['Date'][15])

            # срочная печать данных за интервал, который уже прошел
            if (minNext >= 29 and am_pm == False) or (minNext < 29 and am_pm == True):
                am_pm = not am_pm
                extra = True
                skip = True
                twice_5 = False
                i5 += 1

            # дублирование записей в промежутке времени
            twice = (last_min == min)
            last_min = min

            # повторное вхождение
            if twice_5 and ((min >= 25 and min < 30) or (min >= 55 and min <= 59)):
                extra = True
                # ... дополнение к повтору записей (перехватить лишние записи на границе)
                skip = False
            elif twice_5:
                # дублирование на границе завершено (пора печатать значения)
                twice_5 = False
                extra = True
                skip = True

            # подсчет только "правильных" записей
            if (not GO_IN and not extra):
                if (min % 5 == 0):
                    i5 += 1
                else:
                    ir += 1
            # Место обработки json (сложение для осреднения)
            if not extra:
                # тоталы не изменять
                # округление явно целых величин
                # проверить повторно подсчет среднего значения

                TotalState['SWAP_Used'] += int(data[str(key)]
                                               ['data']['system_SWAP_Used'])
                TotalState['SWAP_Total'] = int(
                    data[str(key)]['data']['system_SWAP_Total'])

                TotalState['RAM_Used'] += int(data[str(key)]
                                              ['data']['system_RAM_Used'])
                TotalState['RAM_Total'] = int(
                    data[str(key)]['data']['system_RAM_Total'])

                try:
                    TotalState['Processes_Total'] += int(
                        data[str(key)]['data']['system_Processes_Total'])
                    TotalState['Processes_Stopped'] += int(
                        data[str(key)]['data']['system_Processes_Stopped'])
                    TotalState['Processes_Sleeping'] += int(
                        data[str(key)]['data']['system_Processes_Sleeping'])
                    TotalState['Processes_Running'] += int(
                        data[str(key)]['data']['system_Processes_Running'])
                    TotalState['Processes_Zombie'] += int(
                        data[str(key)]['data']['system_Processes_Zombie'])
                except:
                    TotalState['Processes_Total'] += 0
                    TotalState['Processes_Stopped'] += 0
                    TotalState['Processes_Sleeping'] += 0
                    TotalState['Processes_Running'] += 0
                    TotalState['Processes_Zombie'] += 0

                TotalState['system_LA1'] += float(data[str(key)]
                                                  ['data']['system_LA1'])
                TotalState['system_LA5'] += float(data[str(key)]
                                                  ['data']['system_LA5'])
                TotalState['system_LA15'] += float(
                    data[str(key)]['data']['system_LA15'])

                try:
                    TotalState['system_IDLE'] += float(
                        data[str(key)]['data']['system_IDLE'])
                except:
                    TotalState['system_IDLE'] += 100.0

                TotalState['HDD_xvda1_Used'] += int(
                    data[str(key)]['data']['system_HDD_xvda1_Used'])
                TotalState['HDD_xvda1_Total'] = int(
                    data[str(key)]['data']['system_HDD_xvda1_Total'])

                # только для webrobo (дополнительные поля root)
                if (uName == device['webrobo']):
                    TotalState['HDD_vg-root_Used'] += int(
                        data[str(key)]['data']['system_HDD_vg-root_Used'])
                    TotalState['HDD_vg-root_Total'] = int(
                        data[str(key)]['data']['system_HDD_vg-root_Total'])

            # ---------------------------------------------------------------------
            if ((min == 25) or (min == 55)) or extra:  # вывод среднего за 30 минут
                extra = False
                twice_5 = True

                # flag
                if (not skip):
                    continue
                skip = False

                twice_5 = False
                if not GO_IN:
                    if (i5 == 6) and (ir == 0):
                        strErr = "Все в порядке."
                        _addError(hour, min, strErr)  # check
                    elif (i5 + ir == 5):
                        strErr = "Ненормальная работа: остановка или сбои в работе. Записи в нужном числе."
                        _addError(hour, min, strErr)
                    elif (i5 + ir < 5 and i5 + ir > 0):
                        strErr = "Остановка работы на некоторое время. Записи частично утеряны."
                        _addError(hour, min, strErr)
                    elif i5 + ir == 0:  # non work
                        strErr = "Сервер не работал."
                        _addError(hour, min, strErr)

                    if twice:
                        _addError(hour, min, 'Дублирование записей.')
                    elif (i5 + ir > 6):
                        strErr = "Присутствуют лишние записи: возможно, сервер был перезагружен."
                        _addError(hour, min, strErr)
                    twice = False

                # осреднение за 30 минут
                for key in TotalState.keys():
                    if key != 'SWAP_Total' and key != 'RAM_Total' and key != 'HDD_xvda1_Total' and key != 'HDD_vg-root_Total':
                        if (i5 + ir <= 0):
                            TotalState[key] = 0
                        else:
                            TotalState[key] /= (i5 + ir)

                i5 = 0
                ir = 0

                # ----------------------------------------------------------------
                # можно подправить вывод
                book.active.cell(row=5, column=interval).value = float(
                    toFixed(TotalState['SWAP_Used'], 3))
                book.active.cell(row=6, column=interval).value = float(
                    toFixed(TotalState['SWAP_Total'], 3))
                try:
                    book.active.cell(row=7, column=interval).value = TotalState['SWAP_Used'] / TotalState['SWAP_Total']
                except:
                    book.active.cell(row=7, column=interval).value = 0
                book.active.cell(row=7, column=interval).number_format = '0%'

                book.active.cell(row=8, column=interval).value = float(
                    toFixed(TotalState['RAM_Used'], 3))
                book.active.cell(
                    row=9, column=interval).value = TotalState['RAM_Total']
                try:
                    book.active.cell(row=10, column=interval).value = TotalState['RAM_Used'] / TotalState['RAM_Total']
                except:
                    book.active.cell(row=10, column=interval).value = 0
                book.active.cell(row=10, column=interval).number_format = '0%'

                book.active.cell(row=11, column=interval).value = round(
                    float(TotalState['Processes_Total']))
                book.active.cell(row=12, column=interval).value = round(
                    float(TotalState['Processes_Stopped']))
                book.active.cell(row=13, column=interval).value = round(
                    float(TotalState['Processes_Sleeping']))
                book.active.cell(row=14, column=interval).value = round(
                    float(TotalState['Processes_Running']))
                book.active.cell(row=15, column=interval).value = round(
                    float(TotalState['Processes_Zombie']))
                book.active.cell(row=16, column=interval).value = float(
                    toFixed(TotalState['system_LA1'], 3))
                book.active.cell(row=17, column=interval).value = float(
                    toFixed(TotalState['system_LA5'], 3))
                book.active.cell(row=18, column=interval).value = float(
                    toFixed(TotalState['system_LA15'], 3))
                book.active.cell(row=19, column=interval).value = float(
                    toFixed(TotalState['system_IDLE'], 3))
                book.active.cell(row=20, column=interval).value = float(
                    toFixed(TotalState['HDD_xvda1_Used'], 3))
                book.active.cell(
                    row=21, column=interval).value = TotalState['HDD_xvda1_Total']
                try:
                    book.active.cell(
                        row=22, column=interval).value = TotalState['HDD_xvda1_Used'] / TotalState['HDD_xvda1_Total']
                except:
                    book.active.cell(row=22, column=interval).value = 0
                book.active.cell(row=22, column=interval).number_format = '0%'

                if (uName == device['webrobo']):
                    book.active.cell(row=23, column=interval).value = float(
                        toFixed(TotalState['HDD_vg-root_Used'], 3))
                    book.active.cell(
                        row=24, column=interval).value = TotalState['HDD_vg-root_Total']
                    try:
                        book.active.cell(
                            row=25, column=interval).value = TotalState['HDD_vg-root_Used'] / TotalState['HDD_vg-root_Total']
                    except:
                        book.active.cell(row=25, column=interval).value = 0
                    book.active.cell(row=25, column=interval).number_format = '0%'

                interval += 1
                # занулить после вывода
                for strr in TotalState.keys():
                    TotalState[strr] = 0

        GO_IN = False


def Correct():              # поправка значений (преобразование для отображения %)
    for i in range(5, 23):
        book.active['P' + str(i)].value = book.active['O' + str(i)].value
    for i in range(3, 17):
        book.active.cell(row=4, column=i).value = 'V'
        book.active.cell(row=4, column=i).fill = _ok

        swap = book.active.cell(row=7, column=i).value
        ram = book.active.cell(row=10, column=i).value
        hdd = book.active.cell(row=22, column=i).value

        _colorSelected(swap, 7, i)
        _colorSelected(ram, 10, i)
        _colorSelected(hdd, 22, i)

    book.active['P7'].number_format = '0%'
    book.active['P10'].number_format = '0%'
    book.active['P22'].number_format = '0%'


def _colorSelected(_val, _line, _dex):      # метод окраса для назначимого параметра
    if (0 <= _val < 0.65):
        book.active.cell(row=_line, column=_dex).fill = _ok
    elif (0.65 <= _val < 0.85):
        book.active.cell(row=_line, column=_dex).fill = _warn
    elif (0.85 <= _val < 1):
        book.active.cell(row=_line, column=_dex).fill = _err


def addToWeb():             # добавление дополнительных строк для webrobo
    book.active = 1
    book.active.cell(row=23, column=2).value = 'HDD (root) Used'
    book.active.cell(row=23, column=2).font = _bold
    book.active.cell(row=23, column=2).border = _selfbord
    book.active.cell(row=24, column=2).value = 'HDD (root) Total'
    book.active.cell(row=24, column=2).font = _bold
    book.active.cell(row=24, column=2).border = _selfbord
    book.active.cell(row=25, column=2).value = 'HDD (root) %'
    book.active.cell(row=25, column=2).font = _bold
    book.active.cell(row=25, column=2).border = _selfbord
    for i in range(2, 17):
        for j in range(23, 26):
            book.active.cell(row=j, column=i).alignment = Alignment(
                horizontal='center')  # выравнивание по центру
            # граница ячейки
            book.active.cell(row=j, column=i).border = _selfbord
    book.active['P23'].value = book.active['O23'].value
    book.active['P24'].value = book.active['O24'].value
    book.active['P25'].value = book.active['O25'].value
    book.active['P25'].number_format = '0%'

    for i in range(3, 17):
        hddroot = book.active.cell(row=25, column=i).value
        _colorSelected(hddroot, 25, i)


# метод добавления ошибок с соответствующим сообщением
def _addError(s_hour, s_min, str_Err):
    ss_hour = str(s_hour)
    s1_hour = str(s_hour + 1)
    if int(s_hour) < 10:
        ss_hour = '0' + str(s_hour)
        s1_hour = '0' + str(s_hour + 1)
        if int(s_hour) == 9:
            s1_hour = str(s_hour + 1)
    if (s_min >= 25 and s_min <= 30):
        ErrorList.append(ss_hour + ':00 - ' + ss_hour + ':30 -> ' + str_Err)
    elif (s_min >= 55 and s_min <= 59):
        ErrorList.append(ss_hour + ':30 - ' + s1_hour + ':00 -> ' + str_Err)
    return


# вся обработка по открытии файла (.json)
with open(fileName) as json_string:

    # список ошибок (временные промежутки аномалий на сервере)
    ErrorList = ['Time codes of anomaly: ']

    # список названий типов сервера (как в original .json)
    device = {'sev': 'Сервер СЕВ', 'dbrobo': 'Сервер dbrobo',
              'webrobo': 'Сервер webrobo', 'dokuwiki': 'Сервер dokuwiki'}

    # текущее состояние для вывода осредненных значений
    TotalState = {'SWAP_Used': 0, 'SWAP_Total': 0, 'RAM_Used': 0, 'RAM_Total': 0,
                  'Processes_Total': 0, 'Processes_Stopped': 0, 'Processes_Sleeping': 0, 'Processes_Running': 0, 'Processes_Zombie': 0,
                  'system_LA1': 0, 'system_LA5': 0, 'system_LA15': 0, 'system_IDLE': 0,
                  'HDD_xvda1_Used': 0, 'HDD_xvda1_Total': 0,
                  'HDD_vg-root_Used': 0, 'HDD_vg-root_Total': 0}

    strServer = (('dbrobo', '01'), ('webrobo', '01'), ('dokuwiki',
                 '01'), ('sev', '01'), ('sev', '02'), ('sev', '03'))

    # словарь данных
    data = json.load(json_string)
    startValues()

    for strData in strServer:
        startValues()
        Processing(strData[0], strData[1])
        bar.next()
        Correct()
        if (strData[0] == 'webrobo'):
            addToWeb()

    bar.finish()

    # оформление и вывод списка ошибок
    '''
    if (len(ErrorList) > 1):
        ErrorList.append('Конец списка ошибок.')
    else:
        ErrorList.append('Ошибок в работе не выявлено.')

    for list in ErrorList:
        print(list)
    '''


book.save('res.xlsx')
book.close()
