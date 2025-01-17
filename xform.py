import openpyxl
from openpyxl.styles import *
from openpyxl.utils.cell import get_column_letter

# Создание xlsx файла сразу с шаблоном для одной страницы
# ВЫНЕСТИ СТИЛИЗАЦИЮ В ЛОКАЛЬНЫЙ ФАЙЛ (для остальных скриптов - кроме xForm.py)

def create_excel():
    book = openpyxl.Workbook()  # создаём книгу
    book.remove(book.active)  # удаляем созданный по умолчанию лист

    # -- СОЗДАЁМ СТРАНИЦЫ EXCEL --
    sheet_dbrobo = book.create_sheet(u'Сервер dbrobo', 0)  # Будет первой
    sheet_webrobo = book.create_sheet(u'Сервер webrobo', 1)  # Второй
    sheet_dokuwiki = book.create_sheet(u'Сервер dokuwiki', 2)
    sheet_CEB1 = book.create_sheet(u'Сервер СЕВ (01)', 3)
    sheet_CEB2 = book.create_sheet(u'Сервер СЕВ (02)', 4)
    sheet_CEB3 = book.create_sheet(u'Сервер СЕВ (03)', 5)

    for sheet in range(0, 6):
        book.active = sheet
        cur_sheet = book.active  # открытие текущего листа в док-те Excel

        # -- НАСТРОЙКИ РАЗМЕРА ЯЧЕЙКИ --
        for i in range(1, 101):  # Для строк
            cur_sheet.row_dimensions[i].height = 22

        for i in range(3, 17):  # Для столбцов
            # преобразовываем индекс столбца в его букву
            letter = get_column_letter(i)
            cur_sheet.column_dimensions[letter].width = 10

        # Отдлеьно для столбца с названиями
        cur_sheet.column_dimensions['B'].width = 25

        # -- СТИЛИЗАЦИЯ PRESETS --
       
        # Жирый шрифт
        global _bold
        _bold = Font(name='Calibri', bold=True, size=14, color="000000")
        # Линия границы ячейки
        global _selfbord
        _selfbord = Border(left=Side(style='medium'), right=Side(
            style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        # Пустое место, где пересекаются времен. интервалы и состояние
        cur_sheet.merge_cells('B2:B3')
        cur_sheet['B2'].border = _selfbord
        cur_sheet['B3'].border = _selfbord

        # -- РАБОТА С ВРЕМЕННЫМИ ИНТЕРВАЛАМИ --
        hour_tmp = 9
        for i in range(3, 16, 2):  # Работа с часами
            # объединение ячеек
            cur_sheet.merge_cells(
                start_row=2, start_column=i, end_row=2, end_column=i+1)
            cur_sheet.cell(row=2, column=i).alignment = Alignment(
                horizontal='center')  # выравнивание по центру
            cur_sheet.cell(row=2, column=i).value = hour_tmp  # запись часов
            cur_sheet.cell(row=2, column=i).font = _bold  # жирн. шрифт
            # граница ячейки (левая часть)
            cur_sheet.cell(row=2, column=i).border = _selfbord
            # граница ячейки (правая часть)
            cur_sheet.cell(row=2, column=i+1).border = _selfbord
            hour_tmp += 1

        for i in range(3, 17):  # Работа с минутами
            cur_sheet.cell(row=3, column=i).alignment = Alignment(
                horizontal='center')  # выравнивание по центру
            if (i % 2):
                cur_sheet.cell(row=3, column=i).value = '00'  # запись минут
            else:
                cur_sheet.cell(row=3, column=i).value = '30'
            # граница ячейки
            cur_sheet.cell(row=3, column=i).border = _selfbord
            cur_sheet.cell(row=3, column=i).font = _bold  # жирн. шрифт

        # -- ПЕРЕЧИСЛЕНИЕ ОБРАБАТЫВАЕМОЙ НАМИ ИНФОРМАЦИИ --
        rows = [u'Доступен', 'SWAP Used', 'SWAP Total', 'SWAP %',
                'RAM Used', 'RAM Total', 'RAM %',
                'Proc. Total', 'Proc. Stopped', 'Proc. Sleeping', 'Proc. Running', 'Proc. Zombie',
                'LA1', 'LA5', 'LA15', 'IDLE',
                'HDD (xvda1) Used', 'HDD (xvda1) Total', 'HDD (xvda1) %']

        # -- ЗАПИСЬ ИНФ. В СТОЛБЕЦ --
        for i, row in enumerate(rows, 1):
            cell = cur_sheet.cell(row=i+3, column=2)
            cell.value = rows[i-1]
            # выравнивание по центру
            cur_sheet.cell(
                row=i+3, column=2).alignment = Alignment(horizontal='center')
            cur_sheet.cell(row=i+3, column=2).font = _bold  # жирн. шрифт
            # граница ячейки
            cur_sheet.cell(row=i+3, column=2).border = _selfbord

        # -- УСТАНОВКА ГРАНИЦ --
        for i in range(3, 17):
            for j in range(4, 23):
                cur_sheet.cell(row=j, column=i).alignment = Alignment(
                    horizontal='center')  # выравнивание по центру
                # граница ячейки
                cur_sheet.cell(row=j, column=i).border = _selfbord
    return book
